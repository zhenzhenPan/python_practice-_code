#import openpyxl
from openpyxl import load_workbook
wb = load_workbook("SPMS_api.xlsx")
print(wb.sheetnames)
#sheet = wb.get_sheet_by_name("Sheet1")
sheet = wb.worksheets[0]

print(sheet)

print(sheet["C"])    #   <-第C列
print(sheet["4"])    #   <-第4行
print(sheet["C4"].value)    # c4     <-第C4格的值
print(sheet.max_row)    # 5     <-最大行数
print(sheet.max_column)    # 13     <-最大列数
for i in sheet["C"]:
  print(i.value, end=" ")    # c1 c2 c3 c4 c5 c6 c7 c8 c9 c10     <-C列中的所有值
  
